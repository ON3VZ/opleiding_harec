#!/usr/bin/env python3
"""Definitieve volledigheidscontrole: elke Excel-vraag moet in het platform zitten."""
import openpyxl, json, os, warnings
warnings.filterwarnings('ignore')
ROOT='/home/claude/site/harec-leerplatform'; INSPECT='/home/claude/inspect'

excel_to_chaps = {
    'electriciteit': ['h02','h03','h05','h11'],
    'componenten':   ['h03'],
    'schakelingen':  ['h04'],
    'ontvangers':    ['h07'],
    'zenders':       ['h06'],
    'antennes_en_transmissielijnen': ['h08'],
    'propagatie':    ['h09'],
    'metingen':      ['h10'],
    'emcveiligheid': ['h12','h13'],
    'reglementering':['h14','h15'],
}
def excel_texts(naam):
    wb=openpyxl.load_workbook(f'{INSPECT}/{naam}.xlsx',data_only=True); ws=wb.worksheets[0]
    return [str(ws.cell(row=r,column=2).value).strip()[:40] for r in range(2,ws.max_row+1) if ws.cell(row=r,column=2).value and str(ws.cell(row=r,column=2).value).strip()]
def platform_texts(chaps):
    out=[]
    for c in chaps:
        p=f'{ROOT}/data/vragen_{c}.json'
        if os.path.exists(p):
            for q in json.load(open(p))['vragen']: out.append(str(q['vraag']).strip()[:40])
    return out

print(f"{'Excel':32s} {'Excel#':>7s} {'gevonden':>9s} {'status':>10s}")
print('-'*62)
tot_e=0; tot_g=0; alle_ok=True
for naam,chaps in excel_to_chaps.items():
    et=excel_texts(naam); pt=set(platform_texts(chaps))
    gevonden=sum(1 for t in et if t in pt)
    tot_e+=len(et); tot_g+=gevonden
    ontbr=len(et)-gevonden
    if ontbr>0: alle_ok=False
    print(f"{naam:32s} {len(et):7d} {gevonden:9d} {('OK' if ontbr==0 else f'MIST {ontbr}'):>10s}")
print('-'*62)
print(f"{'TOTAAL':32s} {tot_e:7d} {tot_g:9d}")
print("\n✅ ALLE VRAGEN COMPLEET" if alle_ok else "\n⚠ ER ONTBREKEN VRAGEN")
